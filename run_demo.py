import os
import numpy as np
from setproctitle import *
import cv2
import natsort
from run_efficieintNet import Efficient_net
from run_yolo_seg import Yolo_Seg
from run_mrcnn import Mask_RCNN
import openpyxl
from tqdm import tqdm
from copy import deepcopy
from collections import Counter
import model_configs
def get_images_paths(image_path):
    '''

    :param image_path:
    :return:
    '''
    if (os.path.isfile(image_path)):
        return [image_path]
    elif (os.path.isdir(image_path)):
        file_paths = [x for x in os.listdir(image_path)]
        file_paths = natsort.natsorted(file_paths)
        for i, file in enumerate(file_paths):
            if ('.png' in file or '.jpg' in file or '.JPG' in file):
                pass
            else:
                file_paths[i] = None
        temp_indexes = list(np.where(np.array(file_paths) != None)[0])
        file_paths = [os.path.join(image_path, file_paths[x]) for x in temp_indexes]
        if (len(file_paths) > 0):
            return file_paths
        else:
            raise Exception("No valid image files found, please check dir")
    else:
        raise Exception("image_path is not dir or valid image, please check image_path")

class Skin_lesion:
    def __init__(self,ef_configs,yolo_configs,mrcnn_configs,exp=False):

        self.ef_configs = ef_configs
        self.ef_config_names = []
        self.num_ef_models = len(ef_configs)
        self.ef_models, self.ef_weights = self.__load_ef_net(ef_configs=self.ef_configs)

        self.exp = exp
        self.__set_experiment()
        self.yolo_configs = yolo_configs
        self.mrcnn_configs = mrcnn_configs
        self.num_yolo_models = len(yolo_configs)
        self.num_mrcnn_models = len(mrcnn_configs)

        self.is_metric_write = False
        self.default_blank = []
        self.yolo_seg_models = self.__load_yolo(yolo_configs)
        self.mrcnn_models = self.__load_mrcnn(mrcnn_configs)
    def __set_experiment(self):
        if(self.exp):
            self.exp_wb = openpyxl.Workbook()
            # self.exp_wb.remove_sheet(self.exp_wb['Sheet'])
            self.exp_wb.remove(self.exp_wb['Sheet'])
            cnt=0
            self.exp_wb.create_sheet(title="yolo", index=cnt)
            self.exp_wb.create_sheet(title="mrcnn", index=cnt + 1)
            cnt=2
            for i in range(self.num_ef_models):
                # note 2 4 3 5
                self.exp_wb.create_sheet(title=self.ef_config_names[i] + "_ucr", index=cnt)
                # ㅜㅐㅅㄷ 수정 2
                self.exp_wb.create_sheet(title=self.ef_config_names[i] + "_cr", index=self.num_ef_models+cnt)
                cnt+=1
            # note cnt 6
            cnt += self.num_ef_models
            self.exp_wb.create_sheet(title="hard", index=cnt)
            self.exp_wb.create_sheet(title="soft", index=cnt + 1)



    def __load_ef_net(self,ef_configs):
        models = [None for x in range(self.num_ef_models)]
        ef_weights = [None for x in range(self.num_ef_models)]
        for i, config in enumerate(ef_configs):
            models[i] = Efficient_net(Config=config,device=config.device)
            ef_weights[i] = config.weight
            self.ef_config_names.append(config.__name__)
        # NOTE 가중치 정규화, 가중치 합이 1이 아닐시에 정규화를 수행함
        epsilon = 1e-6
        total_weight = sum(ef_weights)
        if (abs(total_weight - 1) < epsilon):
            pass
        else:
            for i in range(len(ef_weights)):
                ef_weights[i] = ef_weights[i] / total_weight
        return models, ef_weights
    def __load_yolo(self,yolo_configs):
        models = [None for x in range(self.num_yolo_models)]
        for i, config in enumerate(yolo_configs):
            models[i] = Yolo_Seg(config=config)
        return models
    def __load_mrcnn(self,mrcnn_configs):
        models = [None for x in range(self.num_mrcnn_models)]
        for i, config in enumerate(mrcnn_configs):
            models[i] = Mask_RCNN(mrcnn_config=config())
        return models

    def __ef_inference_uncropped_image(self, image_info):
        results = [None for x in range(self.num_ef_models)]
        for i, model in enumerate(self.ef_models):
            infeered_result = model.inference(image_info=image_info)
            class_names = []
            class_names += [result[0] for result in infeered_result]
            class_names = list(set(class_names))
            predictions_dict = {}
            for name in class_names:
                predictions_dict[name] = 0.0
            for j, (class_name, probability) in enumerate(infeered_result):
                predictions_dict[class_name] += probability
            results[i] = predictions_dict
        # exit()
        return results
    def __ef_inference_cropped_image(self, cropped_images):
        results = [None for x in range(self.num_ef_models)]
        for i, model in enumerate(self.ef_models):
            temp_result = []
            class_names = []
            # print("__ef_inference_cropped_image",len(cropped_images))
            for img in cropped_images:
                infeered_result = model.inference(image_info=img)
                class_names += [result[0] for result in infeered_result]
                temp_result.append(infeered_result)
            class_names = list(set(class_names))
            predictions_dict = {}
            for name in class_names:
                predictions_dict[name] = 0.0
            divide_value = len(temp_result)
            for infeered_result in temp_result:
                for j, (class_name, probability) in enumerate(infeered_result):
                    predictions_dict[class_name] += probability/divide_value
            predictions_dict = dict(sorted(predictions_dict.items(), key=lambda x: x[1], reverse=True))
            results[i] = predictions_dict
        return results

    def __yolo_seg_inference(self, image_info):
        inference_results = []
        for i, model in enumerate(self.yolo_seg_models):
            inference_results += model.inference(image_info=image_info,visualize=True)
        yolo_inferred_images = []
        yolo_cropped_images = []
        yolo_status = []
        yolo_confidences = []
        for i, yolo_result in enumerate(inference_results):
            inferred_image, cropped_images, status, confidences = yolo_result
            yolo_inferred_images += [inferred_image]
            yolo_cropped_images += [cropped_images]
            yolo_status += [status]
            yolo_confidences += [confidences]

        return yolo_inferred_images, yolo_cropped_images, yolo_status, yolo_confidences

    def __mrcnn_inference(self,image_info):
        inference_results = []
        for i, model in enumerate(self.mrcnn_models):
            inference_results += model.inference(image_info=image_info)
        mrcnn_inferred_images = []
        mrcnn_cropped_images=[]
        mrcnn_status=[]
        mrcnn_confidences=[]
        for i, mrcnn_result in enumerate(inference_results):
            inferred_image, cropped_images, status, confidences = mrcnn_result
            mrcnn_inferred_images += [inferred_image]
            mrcnn_cropped_images += [cropped_images]
            mrcnn_status += [status]
            mrcnn_confidences += [confidences]
        return mrcnn_inferred_images,mrcnn_cropped_images,mrcnn_status,mrcnn_confidences
    def save_exp(self):
        file_name = ""
        # for i in range(self.num_ef_models):
        #     # file_name += self.ef_config_names[i]+" "+ str(round(self.ef_weights[i],3))
        #     # if(i+1<self.num_ef_models):
        #     #     file_name += ", "
        #     file_name+="result"
        # file_name += '.xlsx'
        file_name += 'result.xlsx'
        self.exp_wb.save("./exp_results/"+file_name)
    def __write_metrics_to_excel(self,ws,len_result):
        ws.append([])
        ws.append([])
        temp = ['','True Positive','True Negative','False Positive','False Negative','Precision','Recall','F1-Score','Accuracy','']
        self.default_blank = ['' for x in range(len(temp))]
        temp += ['img_names','Ground Truth']
        for i in range(len_result):
            if( (i)%2 == 0):
                temp.append('predict')
            else:
                temp.append('confidence')
        ws.append(temp)
        return ['',"=COUNTIFS(L:L, \"atopy\", M:M, \"atopy\")","=COUNTIFS(L:L,\"<>Atopy\",M:M,\"<>Atopy\",L:L,\"<>\",M:M,\"<>\")-1", "=COUNTIFS(L:L, \"<>Atopy\", M:M, \"Atopy\")","=COUNTIFS(L:L, \"Atopy\", M:M, \"<>Atopy\")","=B4/(B4+D4)","=B4/(B4+E4)","=2*(F4*G4)/(F4+G4)","=(B4+C4)/(B4+C4+D4+E4)",'']


    def __write_to_sheet(self, ws, img_name, result, is_metric_write):
        # result = [x for pair in result for x in pair]
        if is_metric_write:
            metric = self.__write_metrics_to_excel(ws, len(result))
            ws.append(metric + [img_name, img_name.split("_")[0]] + result)
        else:
            ws.append(self.default_blank + [img_name, img_name.split("_")[0]] + result)


    def __write_excel(self, img_name, yolo_status, yolo_mean,mrcnn_status,mrcnn_mean,ef_uncropped_results,ef_cropped_results,hard_voting_result,soft_voting_result):
        is_metric_write = not self.is_metric_write
        if is_metric_write:
            self.is_metric_write = True

        ws = self.exp_wb['yolo']
        self.__write_to_sheet(ws, img_name, ["atopy" if yolo_status else "normal",yolo_mean], is_metric_write)
        ws = self.exp_wb['mrcnn']
        self.__write_to_sheet(ws, img_name, ["atopy" if mrcnn_status else "normal", mrcnn_mean], is_metric_write)
        cnt = 2
        for i, result in enumerate(ef_uncropped_results):
            ws = self.exp_wb[self.exp_wb.sheetnames[cnt + i]]
            sorted_list = sorted(result.items(), key=lambda x: x[1], reverse=True)
            result = [item for pair in sorted_list for item in pair]
            self.__write_to_sheet(ws, img_name, result, is_metric_write)
        cnt += len(ef_uncropped_results)
        for i, result in enumerate(ef_cropped_results):
            ws = self.exp_wb[self.exp_wb.sheetnames[cnt + i]]
            sorted_list = sorted(result.items(), key=lambda x: x[1], reverse=True)
            result = [item for pair in sorted_list for item in pair]
            self.__write_to_sheet(ws, img_name, result, is_metric_write)
        cnt += len(ef_cropped_results)


        ws = self.exp_wb['hard']
        hard_counting =Counter(hard_voting_result).most_common(n=1)[0]
        self.__write_to_sheet(ws, img_name, [hard_counting[0],hard_counting[1]]+hard_voting_result, is_metric_write)
        ws = self.exp_wb['soft']
        self.__write_to_sheet(ws, img_name, soft_voting_result, is_metric_write)

    def hard_voting(self,yolo_status,mrcnn_status,ef_uncropped_results,ef_cropped_results):
        voting = []
        if(yolo_status):
            voting.append("atopy")
        if(mrcnn_status):
            voting.append("atopy")
        for result in ef_uncropped_results:
            sorted_list = sorted(result.items(), key=lambda x: x[1], reverse=True)
            result = [item for pair in sorted_list for item in pair]
            voting.append(result[0])
        # print("-" * 50)
        for result in ef_cropped_results:
            sorted_list = sorted(result.items(), key=lambda x: x[1], reverse=True)
            result = [item for pair in sorted_list for item in pair]
            if(len(result)>0):
                voting.append(result[0])
            else:
                voting.append("normal")
        # print("hard",voting)
        # print(Counter(voting).most_common(n=1))
        return voting
    def soft_voting(self,yolo_mean,mrcnn_mean,ef_uncropped_results,ef_cropped_results):
        class_names = []
        divide_value = 2+ len(ef_uncropped_results) + len(ef_cropped_results)
        for infeered_result in ef_uncropped_results:
            class_names += list(infeered_result.keys())

        for infeered_result in ef_cropped_results:
            class_names += list(infeered_result.keys())
        predictions_dict = {}

        for name in class_names:
            predictions_dict[name] = 0.0
        predictions_dict['atopy'] = predictions_dict.get('atopy', 0) + yolo_mean/divide_value + mrcnn_mean/divide_value
        for infeered_result in ef_uncropped_results:
            for key, value in infeered_result.items():
                predictions_dict[key] += value/divide_value

        for infeered_result in ef_cropped_results:
            for key, value in infeered_result.items():
                # print(key, value)
                predictions_dict[key] += value / divide_value
        predictions_dict = dict(sorted(predictions_dict.items(), key=lambda x: x[1], reverse=True))
        # print("soft", predictions_dict)
        soft = [item for pair in predictions_dict.items() for item in pair]
        # print("soft", soft)
        return soft


    def inference(self,image_path,display=False):
        img = cv2.imread(image_path)
        # ef_results = self.__ef_inference(img)

        mrcnn_inferred_images, mrcnn_cropped_images,mrcnn_status, mrcnn_confidences = self.__mrcnn_inference(image_info=deepcopy(img))
        if(len(mrcnn_inferred_images)>0):
            mrcnn_inferred_images = mrcnn_inferred_images[0]
            mrcnn_cropped_images = mrcnn_cropped_images[0]
            mrcnn_status = mrcnn_status[0]
            mrcnn_confidences = np.array(mrcnn_confidences[0])

        yolo_inferred_images, yolo_cropped_images, yolo_status, yolo_confidences = self.__yolo_seg_inference(
            image_info=deepcopy(img))

        yolo_inferred_images = yolo_inferred_images[0]
        yolo_cropped_images = yolo_cropped_images[0]
        yolo_status = yolo_status[0]
        yolo_confidences = np.array(yolo_confidences[0])

        # mrcnn_sum = mrcnn_confidences.sum()
        # yolo_sum = yolo_confidences.sum()
        # note 값을 0으로
        # mrcnn_mean = mrcnn_confidences.mean()
        # yolo_mean = yolo_confidences.mean()
        mrcnn_mean = 0
        yolo_mean = 0

        # print(mrcnn_sum,len(mrcnn_confidences),mrcnn_mean,yolo_sum,len(yolo_confidences),yolo_mean)
        if(display):
            cv2.imshow("original",img)
            for i, data in enumerate(yolo_cropped_images):
                cv2.imshow("yolo_"+str(yolo_confidences[i]),data)

            for i, data in enumerate(mrcnn_cropped_images):
                cv2.imshow("mrcnn_" + str(mrcnn_confidences[i]), data)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
        # whole_cropped_images = yolo_cropped_images+mrcnn_cropped_images
        whole_cropped_images = []
        if(len(yolo_cropped_images)>0):
            whole_cropped_images += yolo_cropped_images
        if (len(mrcnn_cropped_images) > 0):
            whole_cropped_images += mrcnn_cropped_images


        ef_uncropped_results = self.__ef_inference_uncropped_image(image_info=deepcopy(img))
        ef_cropped_results = self.__ef_inference_cropped_image(cropped_images=whole_cropped_images)

        hard_voting_result = self.hard_voting(yolo_status,mrcnn_status,ef_uncropped_results,ef_cropped_results)
        soft_voting_result = self.soft_voting(yolo_mean, mrcnn_mean, ef_uncropped_results, ef_cropped_results)
            # for result in i:
            #     print(result)
        #     print("ddd"*23)
        if (self.exp):
            self.__write_excel(img_name=os.path.basename(image_path).split('.')[0], yolo_status=yolo_status,yolo_mean=yolo_mean,mrcnn_status=mrcnn_status,mrcnn_mean=mrcnn_mean,ef_uncropped_results=ef_uncropped_results,ef_cropped_results=ef_cropped_results,hard_voting_result=hard_voting_result,soft_voting_result=soft_voting_result)
        return yolo_inferred_images, soft_voting_result


if __name__ == '__main__':
    setproctitle('lesion')
    # note 0.89 acc 0.45 0.35
    # ef_configs = [model_configs.Cfg_2nd_EffB0_Su_Cls_41, model_configs.Cfg_2nd_EffB0_Ming_Cls_6]
    # ef_configs = [model_configs.Cfg_1st_EffB7_Su_Cls_5,model_configs.Cfg_2nd_EffB0_Su_Cls_41,model_configs.Cfg_2nd_EffB0_Ming_Cls_41,model_configs.Cfg_2nd_EffB0_Ming_Cls_6,model_configs.Cfg_2nd_EffB7_Ming_Cls_6,
    #               model_configs.Cfg_3rd_EffB0_Ming1_Cls_4,model_configs.Cfg_3rd_EffB0_Ming2_Cls_4,model_configs.Cfg_3rd_EffB0_Ming3_Cls_4,model_configs.Cfg_3rd_EffB0_Ming4_Cls_4]
    ef_configs = [model_configs.Cfg_2nd_EffB0_Su_Cls_41, model_configs.Cfg_3rd_EffB0_Ming3_Cls_4]
    # ef_configs = [Config_6_min]
    yolo_configs = [model_configs.Config_yolo]
    mrcnn_configs = []
    skin_lesion = Skin_lesion(ef_configs=ef_configs,yolo_configs=yolo_configs,mrcnn_configs=mrcnn_configs,exp=True)
    image_path = "test_data/atomom_test_images_samples/"

    image_path_list = get_images_paths(image_path)
    output = []
    for i in tqdm(range(len(image_path_list))):
        image_path = image_path_list[i]
        # print(image_path)
        skin_lesion.inference(image_path=image_path)
        # break

    if(skin_lesion.exp):
        skin_lesion.save_exp()
